import io

from django.test import Client, TestCase
from django.urls import reverse
from PIL import Image

from .models import ImportSession, ProductItem


def make_png_bytes() -> bytes:
    buffer = io.BytesIO()
    Image.new("RGB", (16, 16), "white").save(buffer, "PNG")
    return buffer.getvalue()


class ProcessValidationTests(TestCase):
    def setUp(self):
        self.client = Client()

    def test_index_ok(self):
        resp = self.client.get(reverse("scanner:index"))
        self.assertEqual(resp.status_code, 200)

    def test_process_without_images_rejected(self):
        resp = self.client.post(reverse("scanner:process"), {})
        self.assertEqual(resp.status_code, 400)
        self.assertFalse(resp.json()["ok"])

    def test_process_rejects_more_than_max_images(self):
        resp = self.client.post(
            reverse("scanner:process"),
            {"images": [io.BytesIO(make_png_bytes()) for _ in range(5)]},
        )
        self.assertEqual(resp.status_code, 400)

    def test_process_rejects_non_image_content(self):
        resp = self.client.post(
            reverse("scanner:process"),
            {"images": io.BytesIO(b"not an image")},
            content_type="multipart/form-data; boundary=x",
        )
        self.assertIn(resp.status_code, (400, 429))


class ItemCrudTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.session = ImportSession.objects.create(status=ImportSession.Status.READY)

    def test_create_item(self):
        resp = self.client.post(
            reverse("scanner:item_create", args=[self.session.id]),
            {
                "producto": "Paracetamol 500 mg",
                "cantidad": 10,
                "unidad": "TAB",
                "precio_unitario": "0.50",
                "laboratorio": "",
                "lote": "",
                "vencimiento": "2027-03",
            },
        )
        self.assertEqual(resp.status_code, 204)
        self.assertEqual(self.session.items.count(), 1)

    def test_create_item_invalid(self):
        resp = self.client.post(
            reverse("scanner:item_create", args=[self.session.id]),
            {"producto": "", "cantidad": 0, "precio_unitario": "-1"},
        )
        self.assertEqual(resp.status_code, 200)  # formulario con errores
        self.assertEqual(self.session.items.count(), 0)

    def test_update_and_delete_item(self):
        item = ProductItem.objects.create(
            session=self.session, producto="Ibuprofeno", cantidad=5,
            precio_unitario="1.00",
        )
        resp = self.client.post(
            reverse("scanner:item_update", args=[self.session.id, item.id]),
            {
                "producto": "Ibuprofeno 400 mg",
                "cantidad": 8,
                "unidad": "TAB",
                "precio_unitario": "1.20",
                "vencimiento": "",
            },
        )
        self.assertEqual(resp.status_code, 204)
        item.refresh_from_db()
        self.assertEqual(item.cantidad, 8)

        resp = self.client.post(
            reverse("scanner:item_delete", args=[self.session.id, item.id])
        )
        self.assertEqual(resp.status_code, 204)
        self.assertEqual(self.session.items.count(), 0)

    def test_item_from_other_session_not_touchable(self):
        other = ImportSession.objects.create(status=ImportSession.Status.READY)
        item = ProductItem.objects.create(
            session=other, producto="Ajeno", cantidad=1, precio_unitario="1.00"
        )
        resp = self.client.post(
            reverse("scanner:item_update", args=[self.session.id, item.id]),
            {"producto": "x", "cantidad": 1, "precio_unitario": "1"},
        )
        self.assertEqual(resp.status_code, 404)

    def test_edit_form_renders_numbers_unlocalized(self):
        """Los inputs type=number deben recibir punto decimal, no coma (es-pe)."""
        item = ProductItem.objects.create(
            session=self.session,
            producto="Omeprazol",
            cantidad=1200,
            precio_unitario="2.50",
        )
        resp = self.client.get(
            reverse("scanner:item_update_form", args=[self.session.id, item.id])
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'value="2.50"')
        self.assertContains(resp, 'value="1200"')
        self.assertNotContains(resp, 'value="2,50"')


class ExportAndCloseTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.session = ImportSession.objects.create(status=ImportSession.Status.READY)
        ProductItem.objects.create(
            session=self.session,
            producto="Paracetamol 500 mg",
            cantidad=10,
            laboratorio="Genfar",
            lote="B1",
            vencimiento="2027-03",
            unidad="TAB",
            precio_unitario="0.50",
        )

    def test_export_xlsx(self):
        resp = self.client.get(reverse("scanner:export", args=[self.session.id]))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", resp["Content-Disposition"])
        # el contenido es un xlsx válido y legible
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], "Producto")
        self.assertEqual(rows[1][0], "Paracetamol 500 mg")
        self.assertEqual(rows[1][1], 10)
        self.assertAlmostEqual(rows[1][7], 5.0)

    def test_close_session(self):
        resp = self.client.post(reverse("scanner:close", args=[self.session.id]))
        self.assertEqual(resp.status_code, 204)
        self.assertEqual(resp["HX-Redirect"], reverse("scanner:index"))
        self.session.refresh_from_db()
        self.assertEqual(self.session.status, ImportSession.Status.CLOSED)
        # la mesa de trabajo y la exportación dejan de estar disponibles
        self.assertEqual(
            self.client.get(reverse("scanner:workspace", args=[self.session.id])).status_code,
            410,
        )
        self.assertEqual(
            self.client.get(reverse("scanner:export", args=[self.session.id])).status_code,
            410,
        )


class PairingTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.session = ImportSession.objects.create(status=ImportSession.Status.READY)

    def test_wrong_password_rejected(self):
        resp = self.client.post(
            reverse("scanner:pair_desktop"), {"password": "incorrecta"}
        )
        self.assertContains(resp, "Contraseña incorrecta", status_code=200)

    def test_token_single_use(self):
        from .models import PairingToken

        token = PairingToken.create()
        url = reverse("scanner:pair_mobile", args=[token.token])
        resp = self.client.post(url, {"session_id": str(self.session.id)})
        self.assertEqual(resp.status_code, 200)
        # segundo uso: debe fallar
        resp = self.client.post(url, {"session_id": str(self.session.id)})
        self.assertEqual(resp.status_code, 410)

    def test_pair_status_flow(self):
        from .models import PairingToken

        token = PairingToken.create()
        resp = self.client.get(reverse("scanner:pair_status", args=[token.token]))
        self.assertEqual(resp.json()["state"], "waiting")
        self.client.post(
            reverse("scanner:pair_mobile", args=[token.token]),
            {"session_id": str(self.session.id)},
        )
        resp = self.client.get(reverse("scanner:pair_status", args=[token.token]))
        self.assertEqual(resp.json()["state"], "ready")
