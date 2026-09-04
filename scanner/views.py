from __future__ import annotations

import hmac
import io
import json
from datetime import datetime

import qrcode
import qrcode.image.svg
from django.conf import settings
from django.contrib import messages
from django.core.files.base import ContentFile
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_GET, require_POST

from . import events
from .forms import ProductItemForm
from .models import UNIDAD_CHOICES, ImportSession, PairingToken, ProductItem, ScanImage
from .services import LLMError, build_thumbnail, extract_items, verify_image


# -------------------------------------------------------- límite de uso

def _throttle_remaining(request, key: str, limit: int, window: int) -> int:
    """Límite simple por sesión de navegador.

    Devuelve los segundos de espera restantes (0 = permitido).
    """
    now = timezone.now().timestamp()
    stamps = [t for t in request.session.get(key, []) if now - t < window]
    request.session[key] = stamps
    if len(stamps) >= limit:
        return max(1, int(window - (now - stamps[0])))
    return 0


def _throttle_hit(request, key: str) -> None:
    request.session.setdefault(key, []).append(timezone.now().timestamp())
    request.session.modified = True


# ---------------------------------------------------------------- utilidades

def _items_context(session: ImportSession) -> dict:
    items = list(session.items.all())
    return {
        "session": session,
        "items": items,
        "items_count": len(items),
        "items_total": sum((i.precio_total for i in items), start=0),
        # Contexto extra para renderizar el formulario de "Agregar" dentro de
        # <template id="tpl-add-item"> en session.html (modal client-side).
        # Es estático por sesión; barato de incluir aquí.
        "add_form_action": reverse("scanner:item_create", args=[session.id]),
        "add_is_edit": False,
        "add_item": None,
        "add_errors": {},
        "add_values": _EMPTY_ITEM_VALUES,
        "unidades": UNIDAD_CHOICES,
    }


def _no_content_with_triggers(triggers: dict) -> HttpResponse:
    """204 + HX-Trigger: el cliente cierra el modal y refresca la región."""
    response = HttpResponse(status=204)
    response["HX-Trigger"] = json.dumps(triggers)
    return response


_EMPTY_ITEM_VALUES = {
    "producto": "",
    "cantidad": 1,
    "unidad": "UND",
    "precio_unitario": 1,
    "laboratorio": "",
    "lote": "",
    "vencimiento": "",
}


def _item_values(item: ProductItem) -> dict:
    return {
        "producto": item.producto,
        "cantidad": item.cantidad,
        "unidad": item.unidad,
        "precio_unitario": item.precio_unitario,
        "laboratorio": item.laboratorio,
        "lote": item.lote,
        "vencimiento": item.vencimiento,
    }


def _render_item_form(request, session, form_action, *, is_edit, item, errors, values):
    return render(
        request,
        "scanner/partials/item_form.html",
        {
            "session": session,
            "form_action": form_action,
            "is_edit": is_edit,
            "item": item,
            "errors": errors,
            "values": values,
            "unidades": UNIDAD_CHOICES,
        },
    )


def _session_or_unavailable(request, session_id):
    """Devuelve (session, None) o (None, respuesta_renderizada)."""
    session = get_object_or_404(ImportSession, pk=session_id)
    if not session.is_available:
        return None, render(
            request, "scanner/session_unavailable.html", status=410
        )
    return session, None


# ---------------------------------------------------------------- páginas

@ensure_csrf_cookie  # camera.js sube las fotos con fetch + cabecera X-CSRFToken
@require_GET
def index(request):
    """Pantalla principal de captura (móvil)."""
    return render(
        request,
        "scanner/index.html",
        {
            "max_images": settings.SCANNER_MAX_IMAGES,
            "accept_types": ",".join(sorted(settings.SCANNER_ALLOWED_IMAGE_TYPES)),
        },
    )


@require_GET
def session_workspace(request, session_id):
    """Mesa de trabajo de validación (móvil o PC)."""
    session, unavailable = _session_or_unavailable(request, session_id)
    if unavailable:
        return unavailable
    context = _items_context(session)
    context["images"] = list(session.images.all())
    return render(request, "scanner/session.html", context)


@require_GET
def session_state(request, session_id):
    """Parcial con el estado canónico de la validación (refresco HTMX/WS)."""
    session, unavailable = _session_or_unavailable(request, session_id)
    if unavailable:
        return unavailable
    return render(request, "scanner/partials/validation.html", _items_context(session))


# ---------------------------------------------------------------- proceso OCR

@require_POST
def process(request):
    """Recibe hasta 4 fotos (multipart), las envía al LLM y crea la sesión."""
    wait = _throttle_remaining(request, "throttle_process", limit=10, window=600)
    if wait:
        return JsonResponse(
            {
                "ok": False,
                "error": "Has procesado varios comprobantes seguidos. "
                f"Espera {wait} segundos e inténtalo otra vez.",
            },
            status=429,
        )

    files = request.FILES.getlist("images")
    max_mb = settings.SCANNER_MAX_IMAGE_BYTES // (1024 * 1024)

    if not files:
        return JsonResponse(
            {"ok": False, "error": "No se recibió ninguna foto. Toma al menos una."},
            status=400,
        )
    if len(files) > settings.SCANNER_MAX_IMAGES:
        return JsonResponse(
            {
                "ok": False,
                "error": f"Puedes procesar hasta {settings.SCANNER_MAX_IMAGES} fotos a la vez.",
            },
            status=400,
        )
    for f in files:
        if f.content_type not in settings.SCANNER_ALLOWED_IMAGE_TYPES:
            return JsonResponse(
                {
                    "ok": False,
                    "error": "Una de las fotos tiene un formato no válido. "
                    "Usa JPG, PNG o WebP.",
                },
                status=400,
            )
        if f.size > settings.SCANNER_MAX_IMAGE_BYTES:
            return JsonResponse(
                {
                    "ok": False,
                    "error": f"Una de las fotos es demasiado pesada (máx. {max_mb} MB). "
                    "Tómala de nuevo más de cerca.",
                },
                status=400,
            )

    raw_images = [(f.read(), f.content_type) for f in files]
    if not all(verify_image(data) for data, _ in raw_images):
        return JsonResponse(
            {
                "ok": False,
                "error": "Una de las fotos está dañada o no es una imagen real. "
                "Tómala de nuevo.",
            },
            status=400,
        )
    _throttle_hit(request, "throttle_process")

    with transaction.atomic():
        session = ImportSession.objects.create(
            status=ImportSession.Status.PROCESSING
        )
        for position, (data, mimetype) in enumerate(raw_images):
            ext = "jpg" if mimetype == "image/jpeg" else mimetype.split("/")[-1]
            scan = ScanImage(
                session=session,
                position=position,
                image=ContentFile(data, name=f"foto-{position + 1}.{ext}"),
            )
            thumb = build_thumbnail(data, name=f"thumb-{position}.jpg")
            if thumb is not None:
                scan.thumb = thumb
            scan.save()

    events.broadcast(session.id, "processing_started", {"photos": len(raw_images)})

    try:
        extracted = extract_items(raw_images)
    except LLMError as exc:
        events.broadcast(session.id, "processing_error", {"message": str(exc)})
        # La sesión fallida no sirve; se elimina junto con sus fotos.
        for img in session.images.all():
            img.image.delete(save=False)
            if img.thumb:
                img.thumb.delete(save=False)
        session.delete()
        return JsonResponse({"ok": False, "error": str(exc)}, status=502)

    with transaction.atomic():
        ProductItem.objects.bulk_create(
            [ProductItem(session=session, **item) for item in extracted]
        )
        session.status = ImportSession.Status.READY
        session.save(update_fields=["status", "updated_at"])

    events.broadcast(
        session.id, "processing_completed", {"items_count": len(extracted)}
    )
    return JsonResponse(
        {
            "ok": True,
            "session_url": reverse("scanner:workspace", args=[session.id]),
        }
    )


# ---------------------------------------------------------------- ítems

@require_GET
def item_create_form(request, session_id):
    session, unavailable = _session_or_unavailable(request, session_id)
    if unavailable:
        return unavailable
    return _render_item_form(
        request,
        session,
        reverse("scanner:item_create", args=[session.id]),
        is_edit=False,
        item=None,
        errors={},
        values=_EMPTY_ITEM_VALUES,
    )


@require_POST
def item_create(request, session_id):
    session, unavailable = _session_or_unavailable(request, session_id)
    if unavailable:
        return unavailable
    form = ProductItemForm(request.POST)
    if not form.is_valid():
        return _render_item_form(
            request,
            session,
            reverse("scanner:item_create", args=[session.id]),
            is_edit=False,
            item=None,
            errors=form.errors,
            values=request.POST,
        )
    with transaction.atomic():
        item = form.save(commit=False)
        item.session = session
        item.position = (session.items.count())
        item.save()
        session.save(update_fields=["updated_at"])
    events.broadcast(session.id, "item_created", {"item_id": str(item.id)})
    return _no_content_with_triggers(
        {
            "closeModal": True,
            "itemsChanged": True,
            "showToast": {"message": "Producto agregado al detalle", "kind": "ok"},
        }
    )


@require_GET
def item_update_form(request, session_id, item_id):
    session, unavailable = _session_or_unavailable(request, session_id)
    if unavailable:
        return unavailable
    item = get_object_or_404(ProductItem, pk=item_id, session=session)
    return _render_item_form(
        request,
        session,
        reverse("scanner:item_update", args=[session.id, item.id]),
        is_edit=True,
        item=item,
        errors={},
        values=_item_values(item),
    )


@require_POST
def item_update(request, session_id, item_id):
    session, unavailable = _session_or_unavailable(request, session_id)
    if unavailable:
        return unavailable
    item = get_object_or_404(ProductItem, pk=item_id, session=session)
    form = ProductItemForm(request.POST, instance=item)
    if not form.is_valid():
        return _render_item_form(
            request,
            session,
            reverse("scanner:item_update", args=[session.id, item.id]),
            is_edit=True,
            item=item,
            errors=form.errors,
            values=request.POST,
        )
    form.save()
    session.save(update_fields=["updated_at"])
    events.broadcast(session.id, "item_updated", {"item_id": str(item.id)})
    return _no_content_with_triggers(
        {
            "closeModal": True,
            "itemsChanged": True,
            "showToast": {"message": "Cambios guardados", "kind": "ok"},
        }
    )


@require_GET
def item_delete_form(request, session_id, item_id):
    session, unavailable = _session_or_unavailable(request, session_id)
    if unavailable:
        return unavailable
    item = get_object_or_404(ProductItem, pk=item_id, session=session)
    return render(
        request,
        "scanner/partials/item_delete.html",
        {
            "session": session,
            "item": item,
            "form_action": reverse(
                "scanner:item_delete", args=[session.id, item.id]
            ),
        },
    )


@require_POST
def item_delete(request, session_id, item_id):
    session, unavailable = _session_or_unavailable(request, session_id)
    if unavailable:
        return unavailable
    item = get_object_or_404(ProductItem, pk=item_id, session=session)
    nombre = item.producto
    item.delete()
    session.save(update_fields=["updated_at"])
    events.broadcast(session.id, "item_deleted", {"item_id": str(item_id)})
    corto = nombre if len(nombre) <= 28 else nombre[:28] + "…"
    return _no_content_with_triggers(
        {
            "closeModal": True,
            "itemsChanged": True,
            "showToast": {"message": f"Se eliminó «{corto}»", "kind": "warn"},
        }
    )


@require_GET
def pair_help(request, session_id):
    """Modal con instrucciones para continuar en la PC."""
    session, unavailable = _session_or_unavailable(request, session_id)
    if unavailable:
        return unavailable
    pair_url = request.build_absolute_uri(reverse("scanner:pair_desktop"))
    return render(
        request,
        "scanner/partials/pair_help.html",
        {"session": session, "pair_url": pair_url},
    )


# ---------------------------------------------------------------- cierre y exportación

@require_GET
def session_export(request, session_id):
    """Descarga el detalle del comprobante como archivo Excel (.xlsx)."""
    session, unavailable = _session_or_unavailable(request, session_id)
    if unavailable:
        return unavailable

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Comprobante"

    headers = [
        "Producto",
        "Cantidad",
        "Laboratorio",
        "Lote",
        "Vencimiento",
        "UND",
        "P. Unit. (S/)",
        "P. Total (S/)",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in session.items.all():
        venc = ""
        if item.vencimiento and len(item.vencimiento) == 7:
            venc = f"{item.vencimiento[5:7]}/{item.vencimiento[0:4]}"
        ws.append(
            [
                item.producto,
                item.cantidad,
                item.laboratorio,
                item.lote,
                venc,
                item.unidad,
                float(item.precio_unitario),
                float(item.precio_total),
            ]
        )

    total = sum((i.precio_total for i in session.items.all()), start=0)
    ws.append([])
    count = session.items.count()
    ws.append(
        [f"TOTAL: {count} productos", "", "", "", "", "", "", float(total)]
    )
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col, width in zip("ABCDEFGH", (42, 10, 25, 14, 13, 10, 13, 13)):
        ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    ahora = datetime.now()
    ahora_formateado = ahora.strftime("%d-%m-%Y-%H-%M")
    response["Content-Disposition"] = (
        # f'attachment; filename="comprobante-{str(session.id)}.xlsx"'
        f'attachment; filename="comprobante-{ahora_formateado}.xlsx"'
    )
    return response


def session_close(request, session_id):
    """GET: modal de confirmación. POST: cierra la sesión y avisa por WS."""
    session, unavailable = _session_or_unavailable(request, session_id)
    if unavailable:
        return unavailable

    if request.method == "GET":
        return render(
            request,
            "scanner/partials/session_close.html",
            {
                "session": session,
                "form_action": reverse("scanner:close", args=[session.id]),
                "items_count": session.items.count(),
            },
        )

    session.status = ImportSession.Status.CLOSED
    session.save(update_fields=["status", "updated_at"])
    events.broadcast(session.id, "session_closed", {})

    response = HttpResponse(status=204)
    response["HX-Redirect"] = reverse("scanner:index")
    return response


# ---------------------------------------------------------------- emparejamiento QR

def _qr_svg(data: str) -> str:
    img = qrcode.make(data, image_factory=qrcode.image.svg.SvgPathImage, box_size=12)
    buffer = io.BytesIO()
    img.save(buffer)
    svg = buffer.getvalue().decode("utf-8")
    # quita la declaración XML para incrustar el SVG en línea
    return svg.split("?>", 1)[-1]


def pair_desktop(request):
    """Pantalla de la PC: contraseña → QR de emparejamiento."""
    context = {"stage": "password", "error": ""}

    if request.method == "POST":
        wait = _throttle_remaining(
            request, "throttle_pair", limit=5, window=300
        )
        if wait:
            context["error"] = (
                f"Demasiados intentos. Espera {wait} segundos e inténtalo otra vez."
            )
            return render(request, "scanner/pair_desktop.html", context)

        password = request.POST.get("password", "")
        if hmac.compare_digest(password, settings.PAIRING_PASSWORD):
            token = PairingToken.create()
            link = request.build_absolute_uri(
                reverse("scanner:pair_mobile", args=[token.token])
            )
            context = {
                "stage": "qr",
                "qr_svg": _qr_svg(link),
                "link": link,
                "status_url": reverse("scanner:pair_status", args=[token.token]),
                "ttl": settings.PAIRING_TOKEN_TTL_SECONDS,
            }
        else:
            _throttle_hit(request, "throttle_pair")
            context["error"] = "Contraseña incorrecta. Inténtalo de nuevo."

    return render(request, "scanner/pair_desktop.html", context)


@require_GET
def pair_status(request, token):
    """La PC consulta si el móvil ya vinculó la sesión (polling ligero)."""
    try:
        pairing = PairingToken.objects.select_related("session").get(token=token)
    except PairingToken.DoesNotExist:
        return JsonResponse({"state": "expired"})
    if pairing.session_id and pairing.used_at:
        return JsonResponse(
            {
                "state": "ready",
                "session_url": reverse(
                    "scanner:workspace", args=[pairing.session_id]
                ),
            }
        )
    if pairing.is_expired:
        return JsonResponse({"state": "expired"})
    return JsonResponse({"state": "waiting"})


def pair_mobile(request, token):
    """Pantalla del móvil tras escanear el QR: elegir sesión y confirmar."""
    try:
        pairing = PairingToken.objects.get(token=token)
    except PairingToken.DoesNotExist:
        pairing = None

    if pairing is None or pairing.used_at or pairing.is_expired:
        return render(request, "scanner/pair_mobile.html", {"stage": "invalid"}, status=410)

    if request.method == "POST":
        session_id = request.POST.get("session_id", "")
        try:
            session = ImportSession.objects.get(pk=session_id)
        except (ImportSession.DoesNotExist, ValueError):
            session = None
        if session is None or not session.is_available:
            messages.error(request, "Esa sesión ya no está disponible.")
        else:
            # vínculo atómico de un solo uso: si otro dispositivo ya usó
            # el token, este update no afecta ninguna fila
            now = timezone.now()
            bound = PairingToken.objects.filter(
                token=pairing.token,
                used_at__isnull=True,
                expires_at__gt=now,
            ).update(session=session, used_at=now)
            if not bound:
                return render(
                    request,
                    "scanner/pair_mobile.html",
                    {"stage": "invalid"},
                    status=410,
                )
            events.broadcast(session.id, "device_connected", {"via": "pairing"})
            return render(
                request,
                "scanner/pair_mobile.html",
                {"stage": "done", "session": session},
            )

    sessions = [
        s
        for s in ImportSession.objects.prefetch_related("items")[:10]
        if s.is_available
    ][:5]
    return render(
        request,
        "scanner/pair_mobile.html",
        {"stage": "confirm", "token": pairing.token, "sessions": sessions},
    )
